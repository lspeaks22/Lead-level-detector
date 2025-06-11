import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox

def process_excel(file_path):
    try:
        df = pd.read_excel(file_path)  # loads the .xlsx file

        df.columns = df.columns.str.strip().str.replace('\n', ' ').str.title()  # cleans the column names
        df.dropna(how='all', inplace=True)  # removes empty rows

        for col in df.columns:
            if df[col].dtype == 'O':
                df[col].fillna("Missing", inplace=True)
            else:
                df[col].fillna(-1, inplace=True)

        # Flag lead levels
        col = "Lead Level" if "Lead Level" in df.columns else "Lead_Level" if "Lead_Level" in df.columns else None
        if col:
            df["Lead Flag"] = df[col].apply(
                lambda x: "High" if isinstance(x, (int, float)) and x > 3.5 else "Normal"
            )  # Adds a "Lead Flag" column; anything over 3.5 µg/dL is flagged

        output_path = file_path.replace(".xlsx", "_CLEANED.xlsx")
        df.to_excel(output_path, index=False)

        wb = load_workbook(output_path)
        ws = wb.active

        # Bold and center headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for col_cells in ws.columns:
            max_len = max(len(str(cell.value)) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = max_len + 2

        # Highlight high lead levels
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        lead_flag_col = None
        lead_level_col = None

        # Identify which columns are Lead Flag and Lead Level
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Lead Flag":
                lead_flag_col = idx
            if cell.value == "Lead Level":
                lead_level_col = idx

        # Apply fill for high lead rows
        if lead_flag_col and lead_level_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[lead_flag_col - 1].value == "High":
                    row[lead_flag_col - 1].fill = red_fill
                    row[lead_level_col - 1].fill = red_fill

        wb.save(output_path)
        messagebox.showinfo("Success", f"Cleaned file saved:\n{output_path}")

    except PermissionError:
        messagebox.showerror("Error", "Permission denied. Please close the Excel file first.")
    except Exception as e:
        messagebox.showerror("Error", f"Something went wrong:\n{str(e)}")

def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        process_excel(file_path)

# GUI
root = tk.Tk()
root.title("Excel Report Generator")
root.geometry("400x200")

label = tk.Label(root, text="📊 Excel Report Generator", font=("Helvetica", 16))
label.pack(pady=20)

browse_btn = tk.Button(root, text="Select Excel File", command=browse_file, font=("Helvetica", 12))
browse_btn.pack(pady=10)

root.mainloop()

